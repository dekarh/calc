from window import Ui_Form
from datetime import datetime
import openpyxl
import NormalizeFields as norm
import sys
import time
from mysql.connector import MySQLConnection, Error
from python_mysql_dbconfig import read_db_config
from PyQt5.QtCore import QDate, QDateTime


class MainWindowSlots(Ui_Form):
    def chk_calendar(self):             # Отлавливаем несуразности в двух календарях
        if self.calendarWidget_2.selectedDate().toPyDate() < self.calendarWidget.selectedDate().toPyDate():
            self.calendarWidget.setSelectedDate(self.calendarWidget_2.selectedDate())
        if self.calendarWidget.selectedDate().toPyDate() < datetime.now().date():
            self.calendarWidget.setSelectedDate(datetime.now().date())
        if self.calendarWidget_2.selectedDate().toPyDate() < datetime.now().date():
            self.calendarWidget_2.setSelectedDate(datetime.now().date())
        delta = self.calendarWidget_2.selectedDate().toPyDate() - self.calendarWidget.selectedDate().toPyDate()
        delta_in_month = 1 + (self.calendarWidget_2.selectedDate().toPyDate().year -
                              self.calendarWidget.selectedDate().toPyDate().year) * 12 + \
                              self.calendarWidget_2.selectedDate().toPyDate().month - \
                              self.calendarWidget.selectedDate().toPyDate().month
        self.label_5.setText('на '+ str(delta.days) + ' дн. ('+ str(delta_in_month) +' мес.)')

    def calc_from_mysql(self):                                         # Рассчитываем комиссию гарантии из MySQL
        summ = float(norm.normalize_float(str(self.lineEdit.text())))  # Проверяем то, что нам ввели (сумму, календари)
        if summ == 0:
            self.label_4.setText('Сумма должна быть больше 0')
            return None
        delta = self.calendarWidget_2.selectedDate().toPyDate() - self.calendarWidget.selectedDate().toPyDate()
        delta_in_month = 1 + (self.calendarWidget_2.selectedDate().toPyDate().year -
                              self.calendarWidget.selectedDate().toPyDate().year) * 12 + \
                              self.calendarWidget_2.selectedDate().toPyDate().month - \
                              self.calendarWidget.selectedDate().toPyDate().month
        if delta.days <= 0:
            self.label_4.setText('Разница дат должна быть больше 1 дня')
            return None


#  SELECT banks.bank_id, banks.bank_name, banks.type_rasch, banks.per_day, banks.koef_185_fz, gar_banks.delta,
#  gar_banks.summ, gar_banks.perc_fz_44, gar_banks.min_fz_44 FROM gar_banks,banks WHERE
# (gar_banks.bank_id = banks.bank_id) AND (banks.per_day = TRUE) AND (gar_banks.delta > 364) AND
# (gar_banks.summ > 999000) ORDER BY (gar_banks.delta - 364), (gar_banks.summ - 999000)

#       SELECT banks.bank_id, banks.bank_name, banks.type_rasch, banks.per_day, banks.koef_185_fz, gar_banks.delta, gar_banks.summ, gar_banks.perc_fz_44, gar_banks.min_fz_44 FROM gar_banks,banks WHERE (gar_banks.bank_id = banks.bank_id) AND (banks.per_day = FALSE) AND (gar_banks.delta > 364) AND (gar_banks.summ > 999000) ORDER BY (gar_banks.delta - 364), (gar_banks.summ - 999000)

        dbconfig = read_db_config()
        conn = MySQLConnection(**dbconfig)
        cursor = conn.cursor()
        try:
            sql = "SELECT banks.bank_id, banks.bank_name, banks.type_rasch, banks.per_day, banks.koef_185_fz, " \
                  "gar_banks.delta, gar_banks.summ, gar_banks.perc_fz_44, gar_banks.min_fz_44 FROM gar_banks,banks" \
                  " WHERE (gar_banks.bank_id = banks.bank_id) AND (banks.per_day = TRUE) AND (gar_banks.delta >= %s)" \
                  " AND (gar_banks.summ >= %s) ORDER BY (gar_banks.delta - %s), (gar_banks.summ - %s)"
            cursor.execute(sql, (delta.days, summ, delta.days, summ))
            rows = cursor.fetchall()

            b_bank_id = {}
            b_bank_name = {}
            b_type_rasch = {}
            b_per_day = {}
            b_koef_185_fz = {}
            b_delta = {}
            b_summ = {}
            b_perc_fz_44 = {}
            b_min_fz_44 = {}

            for row in rows:
                if b_bank_name.get(row[0]) == None:
                    b_bank_id[row[0]] = row[0]
                    b_bank_name[row[0]] = row[1]
                    b_type_rasch[row[0]] = row[2]
                    b_per_day[row[0]] = row[3]
                    b_koef_185_fz[row[0]] = row[4]
                    b_delta[row[0]] = row[5]
                    b_summ[row[0]] = row[6]
                    b_perc_fz_44[row[0]] = row[7]
                    b_min_fz_44[row[0]] = row[8]
                elif (b_summ.get(row[0]) < summ) and (b_delta[row[0]] < delta.days):
                    b_delta[row[0]] = row[5]
                    b_summ[row[0]] = row[6]
                    b_perc_fz_44[row[0]] = row[7]
                    b_min_fz_44[row[0]] = row[8]

            sql = "SELECT banks.bank_id, banks.bank_name, banks.type_rasch, banks.per_day, banks.koef_185_fz, " \
                  "gar_banks.delta, gar_banks.summ, gar_banks.perc_fz_44, gar_banks.min_fz_44 FROM gar_banks,banks" \
                  " WHERE (gar_banks.bank_id = banks.bank_id) AND (banks.per_day = FALSE) AND (gar_banks.delta >= %s)" \
                  " AND (gar_banks.summ >= %s) ORDER BY (gar_banks.delta - %s), (gar_banks.summ - %s)"

            cursor.execute(sql, (delta_in_month, summ, delta_in_month, summ))
            rows = cursor.fetchall()
            for row in rows:
                if b_bank_name.get(row[0]) == None:
                    b_bank_id[row[0]] = row[0]
                    b_bank_name[row[0]] = row[1]
                    b_type_rasch[row[0]] = row[2]
                    b_per_day[row[0]] = row[3]
                    b_koef_185_fz[row[0]] = row[4]
                    b_delta[row[0]] = row[5]
                    b_summ[row[0]] = row[6]
                    b_perc_fz_44[row[0]] = row[7]
                    b_min_fz_44[row[0]] = row[8]
                elif (b_summ.get(row[0]) < summ) and (b_delta[row[0]] < delta_in_month):
                    b_delta[row[0]] = row[5]
                    b_summ[row[0]] = row[6]
                    b_perc_fz_44[row[0]] = row[7]
                    b_min_fz_44[row[0]] = row[8]

            s_text = ''
            for k, kk in enumerate(b_bank_id):                                   # вывод
                if b_perc_fz_44[kk] != 0:
                    if b_per_day[kk] == 1:
                        period = delta.days
                    else:
                        period = delta_in_month
                    if b_perc_fz_44[kk] < 0:                            # если отрицательный (абсолютный)
                        res = abs(b_perc_fz_44[kk])
                    elif b_type_rasch[kk] == 1:                         # по типам расчета
                        res = summ * b_perc_fz_44[kk]
                    elif b_type_rasch[kk] == 2:                         # АНКОР
                        res = summ * period * b_perc_fz_44[kk]
                    elif b_type_rasch[kk] == 3:                         # КБ «Локо-Банк» (Вар.2)
                        res = (0.035 * summ) + (period * b_perc_fz_44[kk] * 4 / 365)
                    elif b_type_rasch[kk] == 4:                         # Солид
                        res = summ * b_perc_fz_44[kk] + 990
                    else:                                               # 0 - по умолчанию (дни, с учетом t)
                        res = summ * period * b_perc_fz_44[kk] / 365
                    if b_min_fz_44[kk] > res:
                        res = b_min_fz_44[kk]
                    if not self.checkBox.isChecked():
                        res = res * b_koef_185_fz[kk]
                    if res > 100:
                        s_text = s_text + b_bank_name[kk] + ':  ' + '{:-.0f}'.format(res) + ' руб.\n'

            if s_text == '':
                self.label_4.setText('Нет удовлетворяющих расчету предложений')
            else:
                self.label_4.setText('Результаты подбора предложений:\n\n' + s_text)

        except Error as e:
            self.label_4.setText('Ошибка: ', e)

        cursor.close()
        conn.close()
        return None

    def append_bank_from_excel_to_mysql(self):
        wb = openpyxl.load_workbook(filename='calc-v11.xlsx', read_only=True)
        k = 0
        i_string = []
        j_string = []
        min_i_res = []
        mi = []
        min_j_res = []
        mj = []
        i_name = []
        j_name = []
        k_name = []
        for k, sheet in enumerate(wb.worksheets):              # Считываем заголовки таблиц и определяем размеры таблиц
            if k == 0:                                         # Пропускаем первую страницу - сам калькулятор
                k_name.append(sheet.title)
                i_name.append([0])
                j_name.append([0])
                min_i_res.append(mi)
                min_j_res.append(mj)
                continue
            for j, excl_row in enumerate(sheet.rows):
                if j == 0:
                    for i, lj in enumerate(excl_row):
                        if lj.value == None or lj.value == 0: # or i == len(excl_row)-1:
                            break
                        else:
                            i_string.append(lj.value)
                if excl_row[0].value == None:
                    break
                j_string.append(excl_row[0].value)
            k_name.append(sheet.title)
            i_name.append(i_string)
            j_name.append(j_string)
            min_i_res.append(mi)
            min_j_res.append(mj)
            mi = []
            mj = []
            i_string = []
            j_string = []
            i_name[k][0] = 0                                    # !!! Подставляем 0 в первых значениях заголовков
            j_name[k][0] = 0                                    # !!! Подставляем 0 в первых значениях заголовков

        tarif = []                                              # заполняем мультиматрицы тарифов из excel
        tarifm = []
        for k, sheet in enumerate(wb.worksheets):
            max_i = len(i_name[k])
            max_j = len(j_name[k])
            tarif_tek = [[0] * max_j for i in range(max_i)]
            tarif_tekm = [[0] * max_j for i in range(max_i)]
            for j, excl_row in enumerate(sheet.rows):
                if j < len(j_name[k]):
                    for i, xj in enumerate(i_name[k]):
                        if excl_row[i].value == 'None':
                            tarif_tek[i][j] = 0
                        else:
                            tarif_tek[i][j] = (excl_row[i].value)
                elif j >= len(j_name[k]) and j < (len(j_name[k])*2):
                    for i, xj in enumerate(i_name[k]):
                        if excl_row[i].value == 'None':
                            tarif_tekm[i][j-len(j_name[k])] = 0
                        else:
                            tarif_tekm[i][j-len(j_name[k])] = (excl_row[i].value)
                else:
                    break
            tarif.append(tarif_tek)
            tarifm.append(tarif_tekm)

        db_config = read_db_config()                                    # Окрываем базу MySQL
        conn = MySQLConnection(**db_config)
        cursor = conn.cursor()

        ii = 1
        try:
            for k, ko in enumerate(k_name):
                if k == 0:
                    continue
                sql = "INSERT INTO banks(bank_id,bank_name,type_rasch,koef_185_fz) VALUES(%s,%s,%s,%s)"
                cursor.execute(sql,(k, ko, 1, 0))
                for j, jo in enumerate(j_name[k]):
                    if j == 0:
                        continue
                    for i, io in enumerate(i_name[k]):
                        if i == 0:
                            continue
                        sql = "INSERT INTO gar_banks(gar_id,bank_id,delta,summ,perc_fz_44,min_fz_44) VALUES " \
                              "(%s,%s,%s,%s,%s,%s)"
                        cursor.execute(sql, (ii, k, io, jo, tarif[k][i][j], tarifm[k][i][j]))
                        ii += 1
            conn.commit()
            self.label_4.setText('Получилось')
        except Error as e:
            conn.rollback()                                                    # Rollback in case there is any error
            self.label_4.setText('Не получилось')
        finally:
            cursor.close()
            conn.close()
        return None

    def calc_from_excel(self):                                         # Рассчитываем комиссию гарантии
        summ = float(norm.normalize_float(str(self.lineEdit.text())))       # Проверяем то, что нам ввели (сумму, календари)
        if summ == 0:
            self.label_4.setText('Сумма должна быть больше 0')
            return None
        delta = self.calendarWidget_2.selectedDate().toPyDate() - self.calendarWidget.selectedDate().toPyDate()
        if delta.days <= 0:
            self.label_4.setText('Разница дат должна быть больше 1 дня')
            return None
        if self.checkBox.isChecked():
            wb = openpyxl.load_workbook(filename='44-FZ.xlsx', read_only=True)
        else:
            wb = openpyxl.load_workbook(filename='185-FZ.xlsx', read_only=True)
        k = 0
        i_string = []
        j_string = []
        min_i_res = []
        mi = []
        min_j_res = []
        mj = []
        i_name = []
        j_name = []
        k_name = []
        for k, sheet in enumerate(wb.worksheets):                             # Считываем заголовки таблиц и определяем размеры таблиц
            for j, excl_row in enumerate(sheet.rows):
                if j == 0:
                    for i, lj in enumerate(excl_row):
                        if lj.value == None or lj.value == 0 or i == len(excl_row)-1:
                            break
                        else:
                            i_string.append(lj.value)
                if excl_row[0].value == None:
                    minis = 1
                if minis > 0:                                                 # minis: 0-осн.табл 1-мин.знач.х 2 - выход
                    if excl_row[0].value != None and minis == 1:
                        minis += 1
                        for i, gj in enumerate(excl_row):                     # заполняем массив минимальных значений х
                            if gj.value == None:
                                break
                            else:
                                mi.append(gj.value)
                else:
                    j_string.append(excl_row[0].value)
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!отладить len(excl_row)-1
                    mj.append(excl_row[len(excl_row)-1].value)              # заполняем массив минимальных значений y
            if len(mi) != len(i_string):
                self.label_4.setText('Ошибка в заполнении строки минимальных значений \n по шкале времени таблицы тарифов')
                return None
            if len(mj) != len(j_string):
                self.label_4.setText('Ошибка в заполнении строки минимальных значений \n по шкале суммы таблицы тарифов')
                return None
            k_name.append(sheet.title)
            i_name.append(i_string)
            j_name.append(j_string)
            min_i_res.append(mi)
            min_j_res.append(mj)
            mi = []
            mj = []
            i_string = []
            j_string = []
            i_name[k][0] = 0                                    # !!! Подставляем 0 в первых значениях заголовков
            j_name[k][0] = 0                                    # !!! Подставляем 0 в первых значениях заголовков

        tarif = []                                                          # заполняем мультиматрицы тарифов
        for k, sheet in enumerate(wb.worksheets):
            max_i = len(i_name[k])
            max_j = len(j_name[k])
            tarif_tek = [[0] * max_j for i in range(max_i)]
            for j, excl_row in enumerate(sheet.rows):
                if excl_row[0].value == None:
                    break
                for i, xj in enumerate(i_name[k]):
                    if excl_row[i].value == 'None':
                        tarif_tek[i][j] = 0
                    else:
                        tarif_tek[i][j] = (excl_row[i].value)
            tarif.append(tarif_tek)

        i_find = []                                                         # координата х результата
        j_find = []                                                         # координата y результата
        k_find = []                                                         # результат
        for k, sheet in enumerate(wb.worksheets):                           # определяем коэффициент
            i = 1
            for i, rj in enumerate(i_name[k]):
                if delta.days <= int(rj):
                    break
            if delta.days > int(rj) or i == 0:                              # проверяем на вхождение в границы таблицы
                i_find.append(int(0))
            else:
                i_find.append(i)
            for j, sj in enumerate(j_name[k]):
                if summ <= float(sj):
                    break
            if summ > float(sj) or i == 0:                              # проверяем на вхождение в границы таблицы
                j_find.append(int(0))
            else:
                j_find.append(j)
            if (j_find[k] > 0) and (i_find[k] > 0):
                k_find.append(tarif[k][i_find[k]][j_find[k]])
            else:
                k_find.append(float('0'))                                  # 0 если не поместилось в границы таблицы
        s_text =''
        for k, name in enumerate(k_name):                                   # вывод
            if k_find[k] != 0:
                if k_find[k] > 0:
                    res = summ*delta.days*k_find[k]/365
                else:
                    res = abs(k_find[k])
                if min_i_res[k][i_find[k]] > res:
                    res = min_i_res[k][i_find[k]]
                if min_j_res[k][j_find[k]] > res:
                    res = min_j_res[k][j_find[k]]
                s_text = s_text + k_name[k] + ':  ' + '{:-.2f}'.format(res) + ' руб.\n\n'
        if s_text == '':
            self.label_4.setText('Нет удовлетворяющих расчету предложений')
        else:
            self.label_4.setText('Результаты подбора предложений:\n\n' + s_text)
        return None
